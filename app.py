from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, send_file, session
import openpyxl
import os
from datetime import datetime
from werkzeug.utils import secure_filename
from difflib import SequenceMatcher as IndiceCoincidencia
import re
import csv
import boto3
from botocore.client import Config

app = Flask(__name__)
app.secret_key = 'tu_clave_secreta_aqui'
app.config['SESSION_TYPE'] = 'filesystem'

# Configuración
UPLOAD_FOLDER = 'uploads'
MAPA_FILE = 'mapa_preventivos.xlsx'
ALLOWED_EXTENSIONS = {'xlsx'}
DOWNLOAD_FOLDER = 'downloads'
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)
#
# Cloudflare R2 Configuration
R2_ACCOUNT_ID = os.environ.get('R2_ACCOUNT_ID')
R2_ACCESS_KEY_ID = os.environ.get('R2_ACCESS_KEY_ID')
R2_SECRET_ACCESS_KEY = os.environ.get('R2_SECRET_ACCESS_KEY')
R2_BUCKET_NAME = os.environ.get('R2_BUCKET_NAME', 'preventivos-app')
R2_ENDPOINT = os.environ.get('R2_ENDPOINT', f'https://{R2_ACCOUNT_ID}.r2.cloudflarestorage.com')

def obtener_cliente_r2():
    """Obtiene el cliente de Cloudflare R2 configurado."""
    try:
        if not all([R2_ACCOUNT_ID, R2_ACCESS_KEY_ID, R2_SECRET_ACCESS_KEY]):
            print("Error: Faltan credenciales de R2 en variables de entorno")
            return None
        
        s3_client = boto3.client(
            's3',
            endpoint_url=R2_ENDPOINT,
            aws_access_key_id=R2_ACCESS_KEY_ID,
            aws_secret_access_key=R2_SECRET_ACCESS_KEY,
            region_name='auto'
        )
        return s3_client
    except Exception as e:
        print(f"Error al configurar cliente R2: {str(e)}")
        return None

def subir_archivo_r2(ruta_archivo, nombre_archivo, carpeta_destino):
    """Sube un archivo a Cloudflare R2 en la carpeta especificada."""
    try:
        s3_client = obtener_cliente_r2()
        if not s3_client:
            return False, "No se pudo configurar el cliente R2"
        
        # Construir la clave del objeto (ruta completa)
        clave = f"{carpeta_destino}/{nombre_archivo}"
        
        # Subir el archivo
        s3_client.upload_file(
            ruta_archivo,
            R2_BUCKET_NAME,
            clave,
            ExtraArgs={'ContentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'}
        )
        
        # Generar URL pública (si el bucket es público)
        url_publica = f"{R2_ENDPOINT}/{R2_BUCKET_NAME}/{clave}"
        
        return True, f"Archivo subido correctamente a R2: {url_publica}"
    except Exception as e:
        return False, f"Error al subir a R2: {str(e)}"

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def descargar_archivo_r2(nombre_archivo, carpeta_origen, ruta_destino):
    """Descarga un archivo XLSX desde Cloudflare R2 y lo guarda localmente."""
    try:
        s3_client = obtener_cliente_r2()
        if not s3_client:
            return False, "No se pudo configurar el cliente R2"

        clave = f"{carpeta_origen}/{nombre_archivo}"

        s3_client.download_file(
            R2_BUCKET_NAME,
            clave,
            ruta_destino
        )

        return True, f"Archivo descargado correctamente desde R2: {ruta_destino}"
    except Exception as e:
        return False, f"Error al descargar desde R2: {str(e)}"


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Funciones para comparación de preventivos y emplazamientos
def formatear_emplazamiento(emplazamiento):
    """Formatea el nombre del emplazamiento eliminando patrones comunes."""
    emplazamiento_formateado = re.sub('(2G)', '', emplazamiento)
    emplazamiento_formateado = re.sub('(3G)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(4G)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(C.T.)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(CT)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(REP-INT-GSM)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(--)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(ATW-T)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(ATW-V)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(ATW)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('[-]', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(TSM)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(T.S.M.)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('(E.B.)', '', emplazamiento_formateado)
    emplazamiento_formateado = re.sub('\s?\d', '', emplazamiento_formateado)
    emplazamiento_formateado = emplazamiento_formateado.lstrip()
    return emplazamiento_formateado

def limpiar_title(title):
    """Limpia el title para usar en URL: quita /, dobles espacios y codifica espacios."""
    # Quitar barras
    title = title.replace('/', '')
    # Quitar dobles espacios
    title = re.sub(r'\s+', ' ', title)
    # Codificar espacios como %20
    title = title.replace(' ', '%20')
    return title

def accion_a_tipos(accion):
    """Convierte el texto de acción a tipos abreviados de preventivos."""
    tipos = set()
    accion_upper = accion.upper()
    
    # Mapeo exacto de nombres de preventivos a tipos
    mapeo_tipos = {
        'MP EB BT ANUAL': 'BT',
        'MP EB OC ANUAL': 'OC',
        'MP EB ANTENA ANUAL': 'SA',
        'MP EB ALIM EQ RADIO': 'AE',
        'MP CABLE DE VIDA ANUAL': 'GS',
        'MP EB CF ANUAL': 'CF',
        'MP EB AA ANUAL': 'AA',
        'MP EB ARMARIO INTEMP': 'AI',
        'MP EB BT SBAL ANUAL': 'BT',
        'MP EB CF ANUAL SB': 'CF',
        'MP CABLE DE VIDA SOP. ANT': 'GS',
        'MP EB ANTENA ANUAL PI': 'SA'
    }
    
    # Buscar coincidencias exactas en el texto de acción
    for nombre, tipo in mapeo_tipos.items():
        if nombre in accion_upper:
            tipos.add(tipo)
    
    # Si no se encontraron tipos, devolver lista vacía
    return sorted(list(tipos))

def leer_excel_emplazamientos(archivo):
    """Lee el archivo de emplazamientos y retorna una lista."""
    hoja_excel = archivo.worksheets[0]
    lista = []
    for fila in hoja_excel.rows:
        emplazamiento = fila[4].value
        if emplazamiento:
            emplazamiento_formateado = formatear_emplazamiento(str(emplazamiento))
            latitud = fila[2].value
            altitud = fila[3].value
            lista.append([emplazamiento, emplazamiento_formateado, latitud, altitud])
    return lista

def leer_excel_preventivos(archivo):
    """Lee el archivo de preventivos y retorna una lista."""
    hoja_excel = archivo.worksheets[0]
    lista = []
    cont = 0
    emplazamiento_a = "SALA ELEMENTO"
    accion = ""
    for fila in hoja_excel.rows:
        cont += 1
        if emplazamiento_a != fila[5].value:
            if cont != 2:
                lista.append([emplazamiento_a, emplazamiento_formateado_a, emplazamiento_b,
                             emplazamiento_formateado_b, accion])
            emplazamiento_a = fila[5].value
            emplazamiento_b = fila[6].value
            emplazamiento_formateado_a = formatear_emplazamiento(str(emplazamiento_a)) if emplazamiento_a else ""
            emplazamiento_formateado_b = formatear_emplazamiento(str(emplazamiento_b)) if emplazamiento_b else ""
            accion = str(fila[8].value) if fila[8].value else ""
        else:
            if cont != 1:
                accion = accion + "\n" + str(fila[8].value)
    lista.append([emplazamiento_a, emplazamiento_formateado_a, emplazamiento_b,
                 emplazamiento_formateado_b, accion])
    return lista

def buscar_coincidencia(lista_preventivos, lista_emplazamientos, indice):
    """Busca coincidencias entre preventivos y emplazamientos."""
    lista_coincidencia = []
    lista_sin_coincidencia = []
    lista_multiple = []  # Para guardar preventivos con múltiples coincidencias
    for preventivo in lista_preventivos:
        contador_coincidencias = 0
        lista = []
        nombre_preventivo = preventivo[0]
        accion = preventivo[4]
        emplazamiento_coincidencia = []
        for emplazamiento in lista_emplazamientos:
            if IndiceCoincidencia(None, preventivo[indice], emplazamiento[1]).ratio() > 0.7:
                latitud = emplazamiento[2]
                altitud = emplazamiento[3]
                # Limpiar title y obtener tipos
                title_limpio = limpiar_title(nombre_preventivo)
                tipos = accion_a_tipos(accion)
                # Generar URL con parámetros de tipos individuales
                tipos_params = '&'.join([f"{tipo}=1" for tipo in tipos])
                url = f"https://preventivos-rgkk.onrender.com/rellenar?title={title_limpio}"
                if tipos_params:
                    url += f"&{tipos_params}"
                lista.append(["Preventivos", "#FF0000", latitud, altitud, nombre_preventivo, url, "#71b300", emplazamiento[0]])
                contador_coincidencias += 1
                emplazamiento_coincidencia.append(emplazamiento)
        if contador_coincidencias == 0:
            lista_sin_coincidencia.append(preventivo)
        elif contador_coincidencias == 1:
            lista_coincidencia.append(lista[0])
        else:
            # Guardar preventivo con múltiples coincidencias para revisión
            lista_multiple.append({
                'preventivo': preventivo,
                'coincidencias': lista,
                'indice': indice
            })
    lista_resultado = []
    lista_resultado.append(lista_coincidencia)
    lista_resultado.append(lista_sin_coincidencia)
    lista_resultado.append(lista_multiple)
    return lista_resultado

def coincidencias(lista_preventivos, lista_emplazamientos):
    """Realiza la búsqueda de coincidencias en dos vueltas."""
    lista_preventivos = buscar_coincidencia(lista_preventivos, lista_emplazamientos, 1)
    lista_coincidencia = lista_preventivos[0]
    lista_sin_coincidencia = lista_preventivos[1]
    lista_multiple_vuelta1 = lista_preventivos[2]
    
    lista_preventivos = buscar_coincidencia(lista_sin_coincidencia, lista_emplazamientos, 3)
    
    for item in lista_preventivos[0]:
        lista_coincidencia.append(item)
    for item in lista_preventivos[1]:
        # Generar URL para preventivos sin coincidencia
        nombre_preventivo = item[0]
        accion = item[4]
        # Limpiar title y obtener tipos
        title_limpio = limpiar_title(nombre_preventivo)
        tipos = accion_a_tipos(accion)
        # Generar URL con parámetros de tipos individuales
        tipos_params = '&'.join([f"{tipo}=1" for tipo in tipos])
        url = f"https://preventivos-rgkk.onrender.com/rellenar?title={title_limpio}"
        if tipos_params:
            url += f"&{tipos_params}"
        lista_coincidencia.append(["Preventivos", "#FF0000", 0, 0, nombre_preventivo, url, "#FF0000"])
    
    # Combinar las múltiples coincidencias de ambas vueltas
    lista_multiple = lista_multiple_vuelta1 + lista_preventivos[2]
    
    return lista_coincidencia, lista_multiple

def obtener_tipo_preventivo(nombre_archivo):
    """Determina el tipo de preventivo basándose en el nombre del archivo."""
    nombre_upper = nombre_archivo.upper()
    
    if 'ALIMENTACIÓN' in nombre_upper or 'ALIMENTACION' in nombre_upper:
        return 'AE'
    elif 'ARMARIOS' in nombre_upper and 'INTEMPERIE' in nombre_upper:
        return 'AI'
    elif '_EBAA_' in nombre_upper or 'EBAA' in nombre_upper:
        return 'AA'
    elif '_SA_' in nombre_upper:
        return 'SA'
    elif '_GS' in nombre_upper:
        return 'GS'
    elif '_OC' in nombre_upper:
        return 'OC'
    elif 'BT' in nombre_upper:
        return 'BT'
    elif '_CF' in nombre_upper:
        return 'CF'
    else:
        # Fallback
        if 'SA' in nombre_upper:
            return 'SA'
        elif 'GS' in nombre_upper:
            return 'GS'
        elif 'OC' in nombre_upper:
            return 'OC'
        elif 'BT' in nombre_upper:
            return 'BT'
        elif 'CF' in nombre_upper:
            return 'CF'
        return 'DESCONOCIDO'

def validar_archivo_preventivo(ruta_archivo):
    """
    Valida si un archivo es un preventivo válido.
    Un preventivo válido debe tener:
    - Una hoja llamada 'Maestra'
    - La hoja Maestra debe estar protegida
    - Tener celdas combinadas
    """
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        
        # Verificar que tiene hoja Maestra
        if 'Maestra' not in wb.sheetnames:
            return False, "El archivo no contiene una hoja llamada 'Maestra'"
        
        ws_maestra = wb['Maestra']
        
        # Verificar que la hoja Maestra está protegida
        if not ws_maestra.protection.sheet:
            return False, "La hoja 'Maestra' no está protegida"
        
        # Verificar que tiene celdas combinadas
        if len(ws_maestra.merged_cells.ranges) == 0:
            return False, "La hoja 'Maestra' no tiene celdas combinadas"
        
        wb.close()
        return True, "Archivo válido"
        
    except Exception as e:
        return False, f"Error al leer el archivo: {str(e)}"

def buscar_celda_por_texto(ws, texto_buscar):
    """
    Busca una celda que contenga el texto especificado.
    Retorna la información de la celda encontrada o None.
    """
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str) and texto_buscar in cell.value:
                # Obtener la celda a la derecha
                col_derecha = cell.column + 1
                celda_derecha = ws.cell(row=cell.row, column=col_derecha)
                
                return {
                    'texto': texto_buscar,
                    'coord': cell.coordinate,
                    'coord_derecha': celda_derecha.coordinate,
                    'valor_derecha': str(celda_derecha.value) if celda_derecha.value else ''
                }
    return None

def analizar_archivo_excel(ruta_archivo):
    """Analiza un archivo Excel y extrae las celdas editables con sus preguntas."""
    nombre_archivo = os.path.basename(ruta_archivo)
    tipo_preventivo = obtener_tipo_preventivo(nombre_archivo)
    
    celdas_info = []
    info_especial = {
        'elemento': None,
        'hoja_validada': []
    }
    
    # Preguntas especiales que siempre son defecto
    preguntas_especiales = ['DNI/MATRICULA:', 'EMPRESA REALIZACIÓN:', 'TELÉFONO:']
    
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        
        # Determinar qué hojas analizar
        hojas_a_analizar = ['Maestra']
        
        # Si existe Hija1, agregarla
        if 'Hija1' in wb.sheetnames:
            hojas_a_analizar.append('Hija1')
        
        # Buscar ELEMENTO: siempre en Maestra
        ws_maestra = wb['Maestra']
        info_especial['elemento'] = buscar_celda_por_texto(ws_maestra, 'ELEMENTO:')
        
        # Buscar HOJA VALIDADA en Maestra y Hija1
        for nombre_hoja in hojas_a_analizar:
            ws = wb[nombre_hoja]
            hoja_validada_info = buscar_celda_por_texto(ws, 'HOJA VALIDADA')
            if hoja_validada_info:
                hoja_validada_info['hoja'] = nombre_hoja
                info_especial['hoja_validada'].append(hoja_validada_info)
        
        # Analizar celdas editables en las hojas correspondientes
        for nombre_hoja in hojas_a_analizar:
            ws = wb[nombre_hoja]
            
            for row in ws.iter_rows():
                for cell in row:
                    # Verificar si la celda no está bloqueada (editable)
                    if not cell.protection.locked:
                        # Evitar celdas combinadas que no son la esquina superior izquierda
                        es_merge = type(cell).__name__ == 'MergedCell'
                        if es_merge:
                            continue
                        
                        # Verificar si pertenece a un rango de celdas combinadas
                        rango_combinado = None
                        esquina_1_1_coord = None
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                rango_combinado = str(merged_range)
                                min_col, min_row, max_col, max_row = merged_range.bounds
                                esquina_1_1_coord = ws.cell(row=min_row, column=min_col).coordinate
                                break
                        
                        # Determinar la coordenada de la celda de respuesta
                        coord_respuesta = esquina_1_1_coord if esquina_1_1_coord else cell.coordinate
                        
                        # Obtener la pregunta a la izquierda
                        col_respuesta = openpyxl.utils.column_index_from_string(coord_respuesta[0])
                        fila_respuesta = int(coord_respuesta[1:])
                        
                        pregunta_texto = ""
                        pregunta_coord = ""
                        
                        if col_respuesta > 1:
                            celda_izquierda = ws.cell(row=fila_respuesta, column=col_respuesta-1)
                            pregunta_coord = celda_izquierda.coordinate
                            
                            # Verificar si la celda izquierda está en un rango combinado
                            pregunta_esquina_1_1 = None
                            for preg_merged in ws.merged_cells.ranges:
                                if celda_izquierda.coordinate in preg_merged:
                                    preg_min_col, preg_min_row, _, _ = preg_merged.bounds
                                    pregunta_esquina_1_1 = ws.cell(row=preg_min_row, column=preg_min_col)
                                    pregunta_coord = pregunta_esquina_1_1.coordinate
                                    pregunta_texto = str(pregunta_esquina_1_1.value) if pregunta_esquina_1_1.value else ""
                                    break
                            
                            if not pregunta_esquina_1_1:
                                pregunta_texto = str(celda_izquierda.value) if celda_izquierda.value else ""
                        
                        # Verificar si tiene validación de datos
                        es_lista = False
                        lista_valores = ""
                        for dv in ws.data_validations.dataValidation:
                            if cell.coordinate in dv.sqref:
                                if getattr(dv, 'type', '') == 'list':
                                    es_lista = True
                                    formula1 = getattr(dv, 'formula1', '')
                                    if formula1 and formula1 != 'N/A':
                                        # Extraer valores de la fórmula (ej: "C,IG,IR,NP")
                                        lista_valores = formula1.strip('"')
                                break
                        
                        # Determinar si es una pregunta especial
                        config_default = None
                        valor_default = ''
                        for preg_especial in preguntas_especiales:
                            if preg_especial in pregunta_texto:
                                config_default = 'defecto'
                                valor_default = ''  # Se llenará con valores globales
                                break
                        
                        celdas_info.append({
                            'pregunta': pregunta_texto,
                            'coord_celda': cell.coordinate,
                            'coord_respuesta': coord_respuesta,
                            'es_lista': es_lista,
                            'lista_valores': lista_valores,
                            'hoja': nombre_hoja,
                            'config_default': config_default,
                            'valor_default': valor_default
                        })
        
        wb.close()
        
    except Exception as e:
        return None, str(e), None
    
    return celdas_info, None, info_especial

def asegurar_mapa_preventivos(tipo_preventivo):
    """
    Asegura que existe el archivo mapa_preventivos.xlsx y la hoja del tipo de preventivo.
    Si no existen, los crea.
    """
    try:
        # Verificar si existe el archivo
        if not os.path.exists(MAPA_FILE):
            # Crear archivo nuevo
            wb = openpyxl.Workbook()
            # Eliminar hoja por defecto
            if 'Sheet' in wb.sheetnames:
                del wb['Sheet']
        else:
            # Abrir archivo existente
            wb = openpyxl.load_workbook(MAPA_FILE)
        
        # Verificar si existe la hoja del tipo de preventivo
        if tipo_preventivo not in wb.sheetnames:
            ws = wb.create_sheet(title=tipo_preventivo)
            # Crear encabezados (sin columna Es Lista, con Hoja)
            headers = ['Pregunta', 'Posición Celda', 'Posición Respuesta', 'Hoja', 'Configuración', 'Valor']
            for col, header in enumerate(headers, start=1):
                ws.cell(row=1, column=col, value=header)
            
            # Ajustar ancho de columnas
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 10
            ws.column_dimensions['E'].width = 15
            ws.column_dimensions['F'].width = 30
        
        wb.save(MAPA_FILE)
        wb.close()
        return True, None
    except PermissionError:
        return False, "El archivo mapa_preventivos.xlsx está abierto en otro programa. Por favor, ciérralo antes de continuar."
    except Exception as e:
        return False, f"Error al gestionar el archivo mapa_preventivos.xlsx: {str(e)}"

def guardar_configuracion_mapa(tipo_preventivo, configuraciones, info_especial):
    """
    Guarda las configuraciones en el archivo mapa_preventivos.xlsx.
    configuraciones es una lista de diccionarios con:
    - pregunta
    - coord_celda
    - coord_respuesta
    - hoja
    - config (ignorar, defecto, equipo, lista)
    - valor
    info_especial contiene información de ELEMENTO y HOJA VALIDADA
    """
    try:
        wb = openpyxl.load_workbook(MAPA_FILE)
        ws = wb[tipo_preventivo]
        
        # Limpiar datos existentes (mantener encabezados)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col, value=None)
        
        fila = 2
        
        # Primero agregar información especial (ELEMENTO y HOJA VALIDADA)
        if info_especial and info_especial.get('elemento'):
            elem = info_especial['elemento']
            ws.cell(row=fila, column=1, value=elem['texto'])
            ws.cell(row=fila, column=2, value=elem['coord'])
            ws.cell(row=fila, column=3, value=elem['coord_derecha'])
            ws.cell(row=fila, column=4, value='Maestra')
            ws.cell(row=fila, column=5, value='INFO')
            ws.cell(row=fila, column=6, value='')  # No guardar valor
            fila += 1
        
        if info_especial and info_especial.get('hoja_validada'):
            for hv in info_especial['hoja_validada']:
                ws.cell(row=fila, column=1, value=hv['texto'])
                ws.cell(row=fila, column=2, value=hv['coord'])
                ws.cell(row=fila, column=3, value=hv['coord_derecha'])
                ws.cell(row=fila, column=4, value=hv['hoja'])
                ws.cell(row=fila, column=5, value='INFO')
                ws.cell(row=fila, column=6, value='')  # No guardar valor
                fila += 1
        
        # Escribir configuraciones de preguntas
        for config in configuraciones:
            ws.cell(row=fila, column=1, value=config['pregunta'])
            ws.cell(row=fila, column=2, value=config['coord_celda'])
            ws.cell(row=fila, column=3, value=config['coord_respuesta'])
            ws.cell(row=fila, column=4, value=config['hoja'])
            ws.cell(row=fila, column=5, value=config['config'])
            ws.cell(row=fila, column=6, value=config.get('valor', ''))
            fila += 1
        
        wb.save(MAPA_FILE)
        wb.close()
        return True, None
    except PermissionError:
        return False, "El archivo mapa_preventivos.xlsx está abierto en otro programa. Por favor, ciérralo antes de continuar."
    except Exception as e:
        return False, f"Error al guardar en el archivo mapa_preventivos.xlsx: {str(e)}"

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/rellenar')
def rellenar():
    """Ruta para rellenar preventivos basándose en parámetros de URL."""
    title = request.args.get('title', '')
    
    # Obtener tipos de preventivos de los parámetros de URL
    tipos = []
    for key in request.args.keys():
        if key.upper() in ['BT', 'OC', 'SA', 'AE', 'GS', 'CF', 'AA', 'AI']:
            if request.args.get(key) == '1':
                tipos.append(key.upper())
    
    # Leer el archivo mapa_preventivos.xlsx
    try:
        ruta_local = MAPA_FILE  
        descargar_archivo_r2('mapa_preventivos.xlsx', 'preventivos', ruta_local)
        
        # NUEVO: Comprobar tamaño del archivo descargado
        tamano = os.path.getsize(ruta_local)
        print(f"Tamaño del archivo descargado: {tamano} bytes")
        if tamano < 1000:
            # Si pesa muy poco, seguramente descargó un error en texto y no el Excel
            with open(ruta_local, 'r', encoding='utf-8', errors='ignore') as f:
                print("Contenido del archivo descargado:", f.read())

        wb = openpyxl.load_workbook(ruta_local)
    except Exception as e:
        return f"CRASH DETECTADO: {str(e)}"  
    # Para cada tipo, buscar la hoja correspondiente y leer las preguntas
    items_por_tipo = {}  # Para mantener el orden original
    equipos_por_tipo = {}  # Para guardar información de equipos por tipo
    
    for tipo in tipos:
        if tipo in wb.sheetnames:
            hoja = wb[tipo]
            items = []  # Para mantener el orden original
            equipos_info = {}  # Para agrupar preguntas por equipo
            equipos_procesados = set()  # Para evitar duplicados
            
            # Leer filas desde la fila 2 (asumiendo que la fila 1 es el encabezado)
            for fila in hoja.iter_rows(min_row=2, values_only=True):
                pregunta = fila[0]  # Columna A
                if pregunta:  # Si hay pregunta
                    config = fila[4] if len(fila) > 4 else None  # Columna E
                    valor = fila[5] if len(fila) > 5 else None  # Columna F
                    
                    # Detectar si es una configuración de equipo
                    es_equipo = False
                    nombre_equipo = None
                    if config and config.lower().startswith('equipo'):
                        es_equipo = True
                        nombre_equipo = config
                    
                    # Solo incluir preguntas con configuración 'lista', 'rellenar' o 'equipoXX'
                    if config and (config.lower() in ['lista', 'rellenar'] or es_equipo):
                        # Si la configuración es 'lista', leer las siguientes columnas hasta encontrar una vacía
                        lista_valores = []
                        if config and config.lower() == 'lista':
                            col_index = 5  # Empezar desde la columna F
                            while col_index < len(fila) and fila[col_index]:
                                lista_valores.append(fila[col_index])
                                col_index += 1
                        
                        pregunta_info = {
                            'pregunta': pregunta,
                            'config': config,
                            'valor': valor,
                            'lista_valores': lista_valores,
                            'es_equipo': es_equipo,
                            'nombre_equipo': nombre_equipo
                        }
                        
                        # Si es equipo, agrupar por nombre de equipo
                        if es_equipo:
                            if nombre_equipo not in equipos_info:
                                equipos_info[nombre_equipo] = []
                            equipos_info[nombre_equipo].append(pregunta_info)
                            
                            # Si es la primera vez que encontramos este equipo, agregar el desplegable
                            if nombre_equipo not in equipos_procesados:
                                items.append({
                                    'tipo': 'equipo_selector',
                                    'nombre_equipo': nombre_equipo
                                })
                                equipos_procesados.add(nombre_equipo)
                            
                            # Agregar la pregunta
                            items.append({
                                'tipo': 'pregunta',
                                'data': pregunta_info
                            })
                        else:
                            items.append({
                                'tipo': 'pregunta',
                                'data': pregunta_info
                            })
            
            # Para cada equipo, leer la hoja correspondiente y obtener los equipos disponibles
            for nombre_equipo, preguntas_equipo in equipos_info.items():
                if nombre_equipo in wb.sheetnames:
                    hoja_equipo = wb[nombre_equipo]
                    equipos_disponibles = []
                    
                    # Leer filas desde la fila 2 (la fila 1 son los títulos)
                    for fila in hoja_equipo.iter_rows(min_row=2, values_only=True):
                        if fila[0]:  # Si hay datos en la primera columna
                            equipos_disponibles.append(fila)
                    
                    # Obtener los títulos de las columnas (fila 1)
                    titulos = []
                    fila_titulos = list(hoja_equipo.iter_rows(min_row=1, max_row=1, values_only=True))
                    if fila_titulos:
                        for titulo in fila_titulos[0]:
                            if titulo:
                                titulos.append(str(titulo))
                    
                    # Si no hay títulos, usar los valores de la columna F de las preguntas como títulos
                    if not titulos:
                        for pregunta in preguntas_equipo:
                            if pregunta['valor']:
                                titulos.append(str(pregunta['valor']))
                    
                    equipos_por_tipo[nombre_equipo] = {
                        'preguntas': preguntas_equipo,
                        'equipos_disponibles': equipos_disponibles,
                        'titulos': titulos
                    }
            
            items_por_tipo[tipo] = items
    
    wb.close()
    
    return render_template('rellenar.html', 
                         title=title,
                         tipos=tipos,
                         items_por_tipo=items_por_tipo,
                         equipos_por_tipo=equipos_por_tipo)

@app.route('/comparar')
def comparar():
    return render_template('comparar.html')

@app.route('/comparar_upload', methods=['POST'])
def comparar_upload():
    if 'file_preventivos' not in request.files or 'file_emplazamientos' not in request.files:
        flash('Debes subir ambos archivos')
        return redirect(url_for('comparar'))
    
    file_preventivos = request.files['file_preventivos']
    file_emplazamientos = request.files['file_emplazamientos']
    
    if file_preventivos.filename == '' or file_emplazamientos.filename == '':
        flash('Debes subir ambos archivos')
        return redirect(url_for('comparar'))
    
    if file_preventivos and allowed_file(file_preventivos.filename) and file_emplazamientos and allowed_file(file_emplazamientos.filename):
        filename_preventivos = secure_filename(file_preventivos.filename)
        filename_emplazamientos = secure_filename(file_emplazamientos.filename)
        
        filepath_preventivos = os.path.join(app.config['UPLOAD_FOLDER'], filename_preventivos)
        filepath_emplazamientos = os.path.join(app.config['UPLOAD_FOLDER'], filename_emplazamientos)
        
        file_preventivos.save(filepath_preventivos)
        file_emplazamientos.save(filepath_emplazamientos)
        
        try:
            # Leer archivos
            wb_preventivos = openpyxl.load_workbook(filepath_preventivos)
            wb_emplazamientos = openpyxl.load_workbook(filepath_emplazamientos)
            
            # Procesar datos
            lista_emplazamientos = leer_excel_emplazamientos(wb_emplazamientos)
            lista_preventivos = leer_excel_preventivos(wb_preventivos)
            
            # Buscar coincidencias
            lista_resultado, lista_multiple = coincidencias(lista_preventivos, lista_emplazamientos)
            
            # Cerrar archivos
            wb_preventivos.close()
            wb_emplazamientos.close()
            
            # Si hay múltiples coincidencias, mostrar página de selección
            if lista_multiple:
                import json
                import uuid
                session_id = str(uuid.uuid4())
                temp_data = {
                    'lista_resultado': lista_resultado,
                    'lista_multiple': lista_multiple
                }
                temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'temp_session_{session_id}.json')
                with open(temp_filepath, 'w', encoding='utf-8') as f:
                    json.dump(temp_data, f)
                
                # Limpiar archivos temporales
                os.remove(filepath_preventivos)
                os.remove(filepath_emplazamientos)
                
                return render_template('seleccionar_coincidencias.html', 
                                     lista_multiple=lista_multiple,
                                     total_coincidencias=len(lista_resultado),
                                     total_multiple=len(lista_multiple),
                                     session_id=session_id)
            
            # Asegurar que todas las filas tengan exactamente 7 columnas (quitando emplazamiento[0] si existe)
            lista_final = []
            for item in lista_resultado:
                if len(item) > 7:
                    lista_final.append(item[:7])
                else:
                    lista_final.append(item)

            # Si no hay múltiples coincidencias, generar CSV directamente
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            csv_filename = f"preventivos_{timestamp}.csv"
            csv_path = os.path.join(DOWNLOAD_FOLDER, csv_filename)
            
            with open(csv_path, mode="w", newline="", encoding="utf-8") as archivo:
                escritor = csv.writer(archivo)
                escritor.writerow(["Folder name", "Folder color", "Latitude", "Longitude", "Title", "Description", "Color"])
                escritor.writerows(lista_final)
            
            # Limpiar archivos temporales
            os.remove(filepath_preventivos)
            os.remove(filepath_emplazamientos)
            
            return send_file(csv_path, as_attachment=True, download_name=csv_filename)
            
        except Exception as e:
            # Limpiar archivos temporales en caso de error
            if os.path.exists(filepath_preventivos):
                os.remove(filepath_preventivos)
            if os.path.exists(filepath_emplazamientos):
                os.remove(filepath_emplazamientos)
            flash(f'Error al procesar los archivos: {str(e)}')
            return redirect(url_for('comparar'))
    
    flash('Tipo de archivo no permitido. Solo se permiten archivos .xlsx')
    return redirect(url_for('comparar'))

@app.route('/procesar_selecciones', methods=['POST'])
def procesar_selecciones():
    data = request.json
    selecciones = data.get('selecciones', [])
    session_id = data.get('session_id', '')
    
    if not session_id:
        return jsonify({'success': False, 'message': 'No se encontró el ID de sesión.'})
        
    temp_filepath = os.path.join(app.config['UPLOAD_FOLDER'], f'temp_session_{session_id}.json')
    if not os.path.exists(temp_filepath):
        return jsonify({'success': False, 'message': 'La sesión temporal ha expirado o no existe.'})
        
    import json
    with open(temp_filepath, 'r', encoding='utf-8') as f:
        temp_data = json.load(f)
        
    lista_resultado = temp_data.get('lista_resultado', [])
    lista_multiple = temp_data.get('lista_multiple', [])
    
    # Procesar selecciones del usuario
    for sel in selecciones:
        index = sel['index']
        seleccion = sel['seleccion']
        
        if index < len(lista_multiple):
            item = lista_multiple[index]
            coincidencia_seleccionada = item['coincidencias'][seleccion]
            # Remover el último elemento (nombre del emplazamiento) para mantener el formato original
            coincidencia_formateada = coincidencia_seleccionada[:-1]
            lista_resultado.append(coincidencia_formateada)
    
    # Asegurar que todas las filas tengan exactamente 7 columnas (quitando emplazamiento[0] si existe)
    lista_final = []
    for item in lista_resultado:
        if len(item) > 7:
            lista_final.append(item[:7])
        else:
            lista_final.append(item)

    # Generar CSV
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    csv_filename = f"preventivos_{timestamp}.csv"
    csv_path = os.path.join(DOWNLOAD_FOLDER, csv_filename)
    
    with open(csv_path, mode="w", newline="", encoding="utf-8") as archivo:
        escritor = csv.writer(archivo)
        escritor.writerow(["Folder name", "Folder color", "Latitude", "Longitude", "Title", "Description", "Color"])
        escritor.writerows(lista_final)
    
    # Limpiar archivo temporal
    if os.path.exists(temp_filepath):
        os.remove(temp_filepath)
    
    return jsonify({
        'success': True,
        'download_url': f'/download_csv/{csv_filename}'
    })

@app.route('/download_csv/<filename>')
def download_csv(filename):
    csv_path = os.path.join(DOWNLOAD_FOLDER, filename)
    return send_file(csv_path, as_attachment=True, download_name=filename)

@app.route('/guardar_equipo', methods=['POST'])
def guardar_equipo():
    """Guarda un nuevo equipo en la hoja correspondiente del Excel."""
    data = request.json
    nombre_equipo = data.get('nombre_equipo')
    datos_equipo = data.get('datos_equipo', [])
    
    if not nombre_equipo or not datos_equipo:
        return jsonify({'success': False, 'message': 'Faltan datos del equipo'})
    
    try:
        # Abrir el archivo Excel
        wb = openpyxl.load_workbook(MAPA_FILE)
        
        # Verificar si existe la hoja del equipo
        if nombre_equipo not in wb.sheetnames:
            wb.close()
            return jsonify({'success': False, 'message': f'No existe la hoja {nombre_equipo}'})
        
        hoja = wb[nombre_equipo]
        
        # Encontrar la primera fila vacía
        fila_vacia = None
        for fila in hoja.iter_rows(min_row=2):
            if not fila[0].value:  # Si la primera celda está vacía
                fila_vacia = fila[0].row
                break
        
        # Si no hay fila vacía, agregar al final
        if fila_vacia is None:
            fila_vacia = hoja.max_row + 1
        
        # Escribir los datos del equipo
        for col_index, valor in enumerate(datos_equipo):
            hoja.cell(row=fila_vacia, column=col_index + 1, value=valor)
        
        # Guardar el Excel
        wb.save(MAPA_FILE)

        wb.close()
        exito_r2, mensaje_r2 = subir_archivo_r2(MAPA_FILE, MAPA_FILE, "preventivos")
        
        return jsonify({'success': True, 'message': 'Equipo guardado correctamente'})
        
    except PermissionError:
        return jsonify({'success': False, 'message': 'El archivo mapa_preventivos.xlsx está abierto en otro programa. Por favor, ciérralo antes de continuar.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al guardar el equipo: {str(e)}'})

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        flash('No se ha seleccionado ningún archivo')
        return redirect(url_for('index'))
    
    file = request.files['file']
    
    if file.filename == '':
        flash('No se ha seleccionado ningún archivo')
        return redirect(url_for('index'))
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        # Validar archivo
        es_valido, mensaje = validar_archivo_preventivo(filepath)
        if not es_valido:
            os.remove(filepath)
            flash(f'Archivo no válido: {mensaje}')
            return redirect(url_for('index'))
        
        # Analizar archivo
        celdas_info, error, info_especial = analizar_archivo_excel(filepath)
        if error:
            os.remove(filepath)
            flash(f'Error al analizar el archivo: {error}')
            return redirect(url_for('index'))
        
        # Obtener tipo de preventivo
        tipo_preventivo = obtener_tipo_preventivo(filename)
        
        # Asegurar mapa_preventivos.xlsx
        exito, error_msg = asegurar_mapa_preventivos(tipo_preventivo)
        if not exito:
            os.remove(filepath)
            flash(error_msg)
            return redirect(url_for('index'))
        
        # Limpiar archivo temporal
        os.remove(filepath)
        
        # Renderizar página de configuración
        return render_template('configurar.html', 
                             tipo_preventivo=tipo_preventivo,
                             celdas_info=celdas_info,
                             info_especial=info_especial,
                             filename=filename)
    
    flash('Tipo de archivo no permitido. Solo se permiten archivos .xlsx')
    return redirect(url_for('index'))

@app.route('/guardar_config', methods=['POST'])
def guardar_config():
    data = request.json
    tipo_preventivo = data.get('tipo_preventivo')
    configuraciones = data.get('configuraciones', [])
    info_especial = data.get('info_especial', {})
    
    exito, error_msg = guardar_configuracion_mapa(tipo_preventivo, configuraciones, info_especial)
    if exito:
        return jsonify({'success': True, 'message': 'Configuración guardada correctamente'})
    else:
        return jsonify({'success': False, 'message': error_msg})

@app.route('/guardar_respuestas', methods=['POST'])
def guardar_respuestas():
    """Guarda las respuestas del formulario en Cloudflare R2."""
    data = request.json
    title = data.get('title', '')
    tipos = data.get('tipos', [])
    respuestas = data.get('respuestas', {})
    
    if not title or not tipos:
        return jsonify({'success': False, 'message': 'Faltan datos: title o tipos'})
    
    try:
        # Leer mapa_preventivos.xlsx para obtener la estructura original
        wb_mapa = openpyxl.load_workbook(MAPA_FILE)
        
        # Crear nuevo workbook para las respuestas
        wb_respuestas = openpyxl.Workbook()
        # Eliminar la hoja por defecto
        if 'Sheet' in wb_respuestas.sheetnames:
            del wb_respuestas['Sheet']
        
        # Obtener el año actual
        año_actual = datetime.now().year
        
        # Por cada tipo de preventivo, crear una hoja idéntica a la del mapa
        for tipo in tipos:
            if tipo not in wb_mapa.sheetnames:
                continue
            
            # Obtener la hoja original del mapa
            hoja_mapa = wb_mapa[tipo]
            # Crear la hoja correspondiente en el excel de respuestas
            hoja_respuestas = wb_respuestas.create_sheet(title=tipo)
            
            # Copiar absolutamente todas las filas y columnas de la hoja original
            for r_idx, row in enumerate(hoja_mapa.iter_rows(values_only=True), start=1):
                for c_idx, val in enumerate(row, start=1):
                    hoja_respuestas.cell(row=r_idx, column=c_idx, value=val)
            
            # Reconstruir la lista de items exactamente igual que en /rellenar
            # para mapear correctamente los nombres de campos del formulario con las filas reales de Excel.
            items_mapeados = []
            equipos_procesados = set()
            
            for row_idx, fila in enumerate(hoja_mapa.iter_rows(min_row=2, values_only=True), start=2):
                pregunta = fila[0]
                if pregunta:
                    config = fila[4] if len(fila) > 4 else None
                    valor = fila[5] if len(fila) > 5 else None
                    
                    es_equipo = False
                    nombre_equipo = None
                    if config and str(config).lower().startswith('equipo'):
                        es_equipo = True
                        nombre_equipo = config
                    
                    # Solo incluir si la configuración es 'lista', 'rellenar' o 'equipoXX'
                    if config and (str(config).lower() in ['lista', 'rellenar'] or es_equipo):
                        if es_equipo and nombre_equipo not in equipos_procesados:
                            # Añadir selector (que incrementa el loop.index en rellenar.html)
                            items_mapeados.append({
                                'tipo': 'equipo_selector',
                                'nombre_equipo': nombre_equipo
                            })
                            equipos_procesados.add(nombre_equipo)
                        
                        # Añadir la pregunta con su fila real
                        if es_equipo:
                            campo_nombre = f"equipo_{nombre_equipo}_{len(items_mapeados)}"
                        else:
                            campo_nombre = f"{tipo}_{len(items_mapeados)}"
                            
                        items_mapeados.append({
                            'tipo': 'pregunta',
                            'row_idx': row_idx,
                            'config': config,
                            'valor_original': valor,
                            'campo_nombre': campo_nombre
                        })
            
            # Actualizar la columna F (columna 6) con las respuestas correspondientes
            for item in items_mapeados:
                if item['tipo'] == 'pregunta':
                    row_idx = item['row_idx']
                    config = item['config']
                    valor_original = item['valor_original']
                    campo_nombre = item['campo_nombre']
                    
                    # Si config es 'ignorar' o 'defecto' (aunque en teoría no se renderizan, se controla por seguridad)
                    if config and str(config).lower() in ['ignorar', 'defecto']:
                        # Mantener el valor original en columna F
                        hoja_respuestas.cell(row=row_idx, column=6, value=valor_original)
                    else:
                        # Rellenar con la respuesta obtenida del formulario
                        respuesta_usuario = respuestas.get(campo_nombre, '')
                        hoja_respuestas.cell(row=row_idx, column=6, value=respuesta_usuario)
                        # Si la configuración era 'lista', limpiar las columnas subsiguientes (G, H, etc.)
                        if config and str(config).lower() == 'lista':
                            for col_to_clear in range(7, hoja_respuestas.max_column + 1):
                                hoja_respuestas.cell(row=row_idx, column=col_to_clear, value='')
                        
        wb_mapa.close()
        
        # Guardar el archivo Excel temporalmente
        nombre_archivo = f"{title}.xlsx"
        ruta_temporal = os.path.join(DOWNLOAD_FOLDER, nombre_archivo)
        wb_respuestas.save(ruta_temporal)
        wb_respuestas.close()
        
        # Subir a Cloudflare R2
        carpeta_destino = f"preventivos/gipuzcoa/{año_actual}"
        exito_r2, mensaje_r2 = subir_archivo_r2(ruta_temporal, nombre_archivo, carpeta_destino)
        
        # Limpiar archivo temporal
        os.remove(ruta_temporal)
        
        if exito_r2:
            return jsonify({
                'success': True, 
                'message': f'Respuestas guardadas correctamente en Cloudflare R2: {mensaje_r2}'
            })
        else:
            return jsonify({
                'success': False, 
                'message': f'Archivo guardado localmente pero error al subir a R2: {mensaje_r2}'
            })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error al guardar respuestas: {str(e)}'})

@app.route('/subir_preventivo', methods=['GET'])
def subir_preventivo():
    return render_template("subir_preventivo.html")

@app.route('/validar_preventivo', methods=['POST'])
def validar_archivo():
# Coincide con formData.append('file', file)
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No se encontró el archivo en la petición'}), 400
    
    file = request.files['file']
    
    if file.filename == '':
        return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400
        
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)        
        tipo_preventivo = obtener_tipo_preventivo(file.filename)
        if tipo_preventivo != "DESCONOCIDO":
            try:
        # 1. Guardar el archivo subido en el servidor antes de leerlo
                filename_preventivos = secure_filename(file_preventivos.filename)
                filepath_preventivos = os.path.join(app.config['UPLOAD_FOLDER'], filename_preventivos)
                file_preventivos.save(filepath_preventivos)  # <-- ¡Importante: faltaba guardar el archivo!

                # 2. Cargar el archivo de preventivos
                wb_preventivos = openpyxl.load_workbook(filepath_preventivos)

                # 3. Descargar y cargar el mapa de preventivos
                ruta_local = MAPA_FILE  
                descargar_archivo_r2('mapa_preventivos.xlsx', 'preventivos', ruta_local)
                wb_mapa = openpyxl.load_workbook(ruta_local)

            except Exception as e:
                 return f"CRASH DETECTADO: {str(e)}"  

                # 4. Seleccionar la hoja (asegúrate de que tipo_preventivo sea int o str según corresponda)
                hoja_mapa = wb_mapa[tipo_preventivo]

                # 5. Iterar correctamente saltándose la cabecera (fila 0)
                coincidencia_mapa = 0
                for i, fila in enumerate(hoja_mapa.rows):
                    if i == 0:
                        continue  # Salta la cabecera
                        
                    pregunta_mapa = fila[0].value
                    celda_mapa = fila[1].value
                    nombre_hoja_preventivo = fila[3].value
                    hoja_preventivo = wb_preventivos[nombre_hoja_preventivo]
                    pregunta_preventivo = hoja_preventivo[celda_mapa].value
                    if pregunta_mapa != pregunta_preventivo:
                        coincidencia_mapa = 1

                if coincidencia_mapa == 0:
                    # Devuelve éxito y el texto que quieres mostrar
                    return jsonify({
                        'success': True, 
                        'message': 'Preventivo detectado correctamente, el mapa esta bien actualizado'
                    })
        else:
            return jsonify({
                'success': False, 
                'error': f'Tipo de preventivo desconocido: {tipo_preventivo} -- {file.filename}'
            }), 400

    return jsonify({'success': False, 'error': 'Formato incorrecto. Sube un archivo .xlsx'}), 400
if __name__ == '__main__':
    app.run(debug=True)
