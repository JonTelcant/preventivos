import openpyxl
import os
from datetime import datetime

def obtener_tipo_preventivo(nombre_archivo):
    """
    Determina el tipo de preventivo basándose en el nombre del archivo.
    
    Args:
        nombre_archivo: Nombre del archivo Excel
        
    Returns:
        str: Abreviatura del tipo de preventivo (AE, SA, GS, OC, BT, CF)
    """
    nombre_upper = nombre_archivo.upper()
    
    if 'ALIMENTACIÓN' in nombre_upper or 'ALIMENTACION' in nombre_upper:
        return 'AE'
    elif '_SA_' in nombre_upper:
        return 'SA'
    elif '_GS' in nombre_upper:
        return 'GS'
    elif '_OC' in nombre_upper:
        return 'OC'
    elif 'BT' in nombre_upper:
        return 'BT'
    elif '_CF' in nombre_upper:
        return 'CF'
    else:
        # Fallback: intentar deducir del nombre
        if 'SA' in nombre_upper:
            return 'SA'
        elif 'GS' in nombre_upper:
            return 'GS'
        elif 'OC' in nombre_upper:
            return 'OC'
        elif 'BT' in nombre_upper:
            return 'BT'
        elif 'CF' in nombre_upper:
            return 'CF'
        return 'DESCONOCIDO'

def analizar_archivo_excel(ruta_archivo):
    """
    Analiza un archivo Excel de preventivos y extrae información de celdas editables.
    
    Args:
        ruta_archivo: Ruta del archivo Excel a analizar
        
    Returns:
        dict: Diccionario con la información extraída del archivo
    """
    nombre_archivo = os.path.basename(ruta_archivo)
    tipo_preventivo = obtener_tipo_preventivo(nombre_archivo)
    
    resultado = {
        'archivo': nombre_archivo,
        'tipo_preventivo': tipo_preventivo,
        'hojas': {},
        'fecha_analisis': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }
    
    try:
        wb = openpyxl.load_workbook(ruta_archivo, data_only=True)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            info_hoja = {
                'protegida': ws.protection.sheet,
                'rango_usado': ws.dimensions,
                'celdas_combinadas': len(ws.merged_cells.ranges),
                'validaciones_datos': len(ws.data_validations.dataValidation),
                'celdas_editables': []
            }
            
            # Si la hoja no está protegida, no tiene celdas bloqueadas
            if not ws.protection.sheet:
                info_hoja['mensaje'] = "Hoja no protegida - todas las celdas son editables"
                resultado['hojas'][sheet_name] = info_hoja
                continue
            
            # Recorrer todas las celdas para encontrar las editables
            for row in ws.iter_rows():
                for cell in row:
                    # Verificar si la celda no está bloqueada (editable)
                    if not cell.protection.locked:
                        # Evitar celdas combinadas que no son la esquina superior izquierda
                        es_merge = type(cell).__name__ == 'MergedCell'
                        if es_merge:
                            continue
                        
                        info_celda = {
                            'coordenada': cell.coordinate,
                            'valor': cell.value,
                            'fila': cell.row,
                            'columna': cell.column
                        }
                        
                        # Verificar si pertenece a un rango de celdas combinadas
                        rango_combinado = None
                        esquina_1_1_coord = None
                        for merged_range in ws.merged_cells.ranges:
                            if cell.coordinate in merged_range:
                                rango_combinado = str(merged_range)
                                min_col, min_row, max_col, max_row = merged_range.bounds
                                esquina_1_1_coord = ws.cell(row=min_row, column=min_col).coordinate
                                info_celda['rango_combinado'] = rango_combinado
                                info_celda['esquina_1_1'] = esquina_1_1_coord
                                info_celda['valor_esquina_1_1'] = ws.cell(row=min_row, column=min_col).value
                                break
                        
                        # Determinar la coordenada de la celda de respuesta (esquina 1:1 si está en rango)
                        coord_respuesta = esquina_1_1_coord if esquina_1_1_coord else cell.coordinate
                        
                        # Obtener la pregunta a la izquierda de la celda de respuesta
                        col_respuesta = openpyxl.utils.column_index_from_string(coord_respuesta[0])
                        fila_respuesta = int(coord_respuesta[1:])
                        
                        pregunta_texto = ""
                        pregunta_coord = ""
                        
                        if col_respuesta > 1:
                            # Obtener celda a la izquierda
                            celda_izquierda = ws.cell(row=fila_respuesta, column=col_respuesta-1)
                            pregunta_coord = celda_izquierda.coordinate
                            
                            # Verificar si la celda izquierda está en un rango combinado
                            pregunta_esquina_1_1 = None
                            for preg_merged in ws.merged_cells.ranges:
                                if celda_izquierda.coordinate in preg_merged:
                                    preg_min_col, preg_min_row, _, _ = preg_merged.bounds
                                    pregunta_esquina_1_1 = ws.cell(row=preg_min_row, column=preg_min_col)
                                    pregunta_coord = pregunta_esquina_1_1.coordinate
                                    pregunta_texto = str(pregunta_esquina_1_1.value) if pregunta_esquina_1_1.value else ""
                                    break
                            
                            if not pregunta_esquina_1_1:
                                pregunta_texto = str(celda_izquierda.value) if celda_izquierda.value else ""
                        
                        info_celda['pregunta_texto'] = pregunta_texto
                        info_celda['pregunta_coord'] = pregunta_coord
                        info_celda['coord_respuesta'] = coord_respuesta
                        
                        # Verificar si tiene validación de datos
                        for dv in ws.data_validations.dataValidation:
                            if cell.coordinate in dv.sqref:
                                info_celda['validacion'] = {
                                    'tipo': getattr(dv, 'type', 'N/A'),
                                    'formula1': getattr(dv, 'formula1', 'N/A'),
                                    'formula2': getattr(dv, 'formula2', 'N/A'),
                                    'operador': getattr(dv, 'operator', 'N/A'),
                                    'showDropDown': getattr(dv, 'showDropDown', 'N/A')
                                }
                                break
                        
                        info_hoja['celdas_editables'].append(info_celda)
            
            resultado['hojas'][sheet_name] = info_hoja
        
        wb.close()
        
    except Exception as e:
        resultado['error'] = str(e)
    
    return resultado

def generar_excel_mapa(resultados, ruta_salida):
    """
    Genera un archivo Excel con el mapa de preventivos, organizado por tipo.
    
    Args:
        resultados: Lista de diccionarios con los resultados de análisis
        ruta_salida: Ruta donde guardar el archivo Excel
    """
    wb_salida = openpyxl.Workbook()
    
    # Eliminar la hoja por defecto
    if 'Sheet' in wb_salida.sheetnames:
        del wb_salida['Sheet']
    
    # Agrupar resultados por tipo de preventivo
    por_tipo = {}
    for resultado in resultados:
        if 'error' in resultado:
            continue
        
        tipo = resultado['tipo_preventivo']
        if tipo not in por_tipo:
            por_tipo[tipo] = []
        
        # Extraer todas las celdas editables de todas las hojas
        for nombre_hoja, info_hoja in resultado['hojas'].items():
            for celda in info_hoja['celdas_editables']:
                por_tipo[tipo].append({
                    'archivo': resultado['archivo'],
                    'hoja': nombre_hoja,
                    'pregunta': celda.get('pregunta_texto', ''),
                    'coord_celda': celda['coordenada'],
                    'coord_respuesta': celda.get('coord_respuesta', celda['coordenada']),
                    'es_lista': 'Sí' if celda.get('validacion', {}).get('tipo') == 'list' else 'No'
                })
    
    # Crear una hoja por cada tipo de preventivo
    for tipo in ['AE', 'SA', 'GS', 'OC', 'BT', 'CF']:
        if tipo not in por_tipo or not por_tipo[tipo]:
            continue
        
        ws = wb_salida.create_sheet(title=tipo)
        
        # Encabezados
        headers = ['Pregunta', 'Posición Celda', 'Posición Respuesta', 'Es Lista']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        fila = 2
        for item in por_tipo[tipo]:
            ws.cell(row=fila, column=1, value=item['pregunta'])
            ws.cell(row=fila, column=2, value=item['coord_celda'])
            ws.cell(row=fila, column=3, value=item['coord_respuesta'])
            ws.cell(row=fila, column=4, value=item['es_lista'])
            fila += 1
        
        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 50
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 10
    
    # Guardar el archivo
    wb_salida.save(ruta_salida)
    wb_salida.close()
    
    print(f"Archivo Excel generado en: {ruta_salida}")
    print(f"Hojas creadas: {', '.join(por_tipo.keys())}")

def main():
    # Lista de archivos a analizar
    archivos = [
        r'c:\Users\etxea\Downloads\MV_Alimentación_Equipo_Radio_F2021_12-PRE-00363845.xlsx',
        r'c:\Users\etxea\Downloads\MV_EB2017_SA_1-PRE-00366049.xlsx',
        r'c:\Users\etxea\Downloads\MV_EB2016_GS-PRE-00427320.xlsx',
        r'c:\Users\etxea\Downloads\MV_EB2014_OC1-PRE-00363218.xlsx',
        r'c:\Users\etxea\Downloads\MV_EBBT_M2025-PRE-00363404.xlsx',
        r'c:\Users\etxea\Downloads\MV_EB2014_CF-PRE-00363435.xlsx'
    ]
    
    print("Iniciando análisis de archivos Excel de preventivos...")
    print("="*60)
    
    resultados = []
    
    for archivo in archivos:
        if os.path.exists(archivo):
            print(f"\nAnalizando: {os.path.basename(archivo)}")
            resultado = analizar_archivo_excel(archivo)
            resultados.append(resultado)
            
            # Resumen rápido
            if 'error' not in resultado:
                total_celdas = sum(len(h['celdas_editables']) for h in resultado['hojas'].values())
                print(f"  ✓ Analizado - {len(resultado['hojas'])} hojas, {total_celdas} celdas editables")
            else:
                print(f"  ✗ Error: {resultado['error']}")
        else:
            print(f"\n⚠ Archivo no encontrado: {archivo}")
    
    # Generar reporte HTML
    ruta_reporte = r'c:\Users\etxea\Documents\Programacion\Python\Preventivos\reporte_analisis.html'
    generar_reporte_html(resultados, ruta_reporte)
    
    print("\n" + "="*60)
    print("Análisis completado.")
    print(f"Total archivos analizados: {len(resultados)}")
    print(f"Reporte generado: {ruta_reporte}")

if __name__ == "__main__":
    main()
